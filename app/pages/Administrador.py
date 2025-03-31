import os
from google.cloud import bigquery
import pandas as pd
import numpy as np
import streamlit as st
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.styles.numbers import BUILTIN_FORMATS
from io import BytesIO
from Funcoes import ler_tabela, incluir_servico, apagar_tabela, incluir_login, alterar_senha, excluir_login
import json
from google.oauth2 import service_account

# os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = "C:\PROJETOS\Pagamento Terceirizado\Ignorar\pagamento-terceirizado-467d410b51b5.json"

# Carrega a chave do Streamlit Secrets
gcp_info = json.loads(st.secrets["gcp_service_account"])

# Cria credencial a partir do dicionário
credentials = service_account.Credentials.from_service_account_info(gcp_info)

# Função para salvar a tabela em um único Excel com formatação
def salvar_excel_com_formatacao(bd):
    output = BytesIO()
    #=== Salvar em uma planilha em excel ===#
    # Crie uma nova planilha Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Servico Prestado"

    # # Remover as linhas de grade do Excel
    # ws.sheet_view.showGridLines = False

    # Define o estilo de preenchimento para o fundo do cabeçalho
    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Define o alinhamento centralizado sem quebra de texto
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

    # Define o alinhamento centralizado
    center_alignment = Alignment(horizontal="center", vertical="center")

    # Define o tamanho da fonte para todas as células
    font_size = Font(size=10)

    # Define a fonte em negrito
    bold_font = Font(bold=True, size=10)

    # Loop pelos DataFrames e escrevendo na planilha
    row_offset = 1  # Inicializa a contagem de linhas na planilha

    # Converter o DataFrame para linhas que o openpyxl pode usar
    rows = dataframe_to_rows(bd, index=False, header=True)

    # Define a borda fina
    borda_fina = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )

    bd["TOTAL"] = pd.to_numeric(bd["TOTAL"], errors="coerce")
    
    # Escreve os dados no Excel
    for j, row in enumerate(rows):
        ws.append(row)

        # Itera sobre todas as colunas e aplica centralização
        for col in range(1, len(row) + 1):  # Itera sobre as colunas
            cell = ws.cell(row=row_offset + j, column=col)
            cell.alignment = center_alignment  # Aplica alinhamento centralizado
            cell.font = font_size # Aplica o tamanho da fonte de 10 para todas as células
            cell.border = borda_fina  # adiciona borda a todas as células

            # Estiliza apenas o cabeçalho
            if j == 0:  # Cabeçalhos do multiíndice
                cell.fill = header_fill
                cell.font = bold_font
                cell.alignment = header_alignment # Estiliza o cabeçalho com quebra de texto

        # Formata como porcentagem colunas 6 e 7
        if (j > 0):  # Dados (exclui os cabeçalhos)
            for col in range(6, 9):  # Colunas 6 e 7
                cell = ws.cell(row=row_offset + j, column=col)
                if isinstance(cell.value, (int, float)):  # Verifica se é número
                    cell.number_format = 'R$ #,##0.00'  # Formato moeda brasileira (R$)
                    cell.value = float(cell.value)  # Converte o valor para floats
    
    # Salvar o Workbook no buffer
    wb.save(output)
    return output.getvalue()


# CSS personalizado
st.markdown(
    """
    <style>
    /* Cor de fundo da página */
    [data-testid="stAppViewContainer"] {
        background-color: #000000;
    }

    /* Cor de fundo do cabeçalho */
    [data-testid="stHeader"] {
        background-color: #000000;
    }

    /* Esconde o menu lateral */
    [data-testid="stSidebar"] {
        display: none;  /* 👈 Esconde o menu lateral */
    }

    /* Remove o espaço lateral */
    [data-testid="stAppViewContainer"] > .main {
        margin-left: 0;  /* 👈 Remove o espaço lateral */
    }

    /* Cor de fundo da barra lateral */
    [data-testid="stSidebar"] {
        background-color: #333333;
    }

    /* Cor do título */
    h1 {
    color: white !important;
    text-align: center;
    font-weight: bold;
}

    /* Cor do subtítulo */
    h2 {
        color: #FFD700;
    }

    /* Cor do texto normal */
    p, span {
        color: #FFFFFF;
    }

    /* Cor dos botões */
    button {
        background-color: #20541B !important;
        color: white !important;
    }

    /* Caixa do formulário */
    div[data-testid="stForm"] {
        background-color: #1e1e1e;  /* cinza escuro */
        padding: 30px;
        border-radius: 12px;
        border: 1px solid #444444;
        max-width: 600px;
        margin: auto;
    }

    /* Campos de texto */
    input, select, textarea {
        /* background-color: #2e2e2e !important; */
        /* color: white !important; */
        border: none !important;
        border-radius: 6px !important;
    }

    /* Botão */
    button[kind="primary"] {
        background-color: #20541B !important;
        color: white !important;
        border-radius: 8px !important;
    }

    table {
    background-color: #000000;
    color: white;
    border-collapse: collapse;
    width: 100%;
    border-radius: 10px;
    overflow: hidden;
    font-size: 14px;
    }
    th, td {
        border: 1px solid #333;
        padding: 8px;
        text-align: left;
    }
    th {
        background-color: #111111;
        color: #FFFFFF;
    }
    tr:nth-child(even) {
        background-color: #1c1c1c;
    }
    </style>
    """,
    unsafe_allow_html=True
)

if "login_admin_sucesso" not in st.session_state or not st.session_state.login_admin_sucesso:
    st.warning("❌ Você precisa fazer login de administrador!")
    st.stop()

#=== Título ===#
st.title("Pagamento Terceirizado")
st.write("")
st.write(f"Bem-vindo, **{st.session_state.LOGIN}**! 😊")
st.write("")
st.write("📋 Tabela com as informações de Serviços que foram prestados pelos colaboradores")

df = ler_tabela(project_id="pagamento-terceirizado", 
                dataset_id="pagamento_terceirizado", 
                table_id="horas_colaborador")
# st.dataframe(df, width=2000, hide_index=True)
# Exibe a tabela com HTML estilizado
st.markdown(df.to_html(index=False, escape=False), unsafe_allow_html=True)

# Link para download
excel_data = salvar_excel_com_formatacao(df)
st.download_button(
    label="Baixar em Excel",
    data=excel_data,
    file_name="Servicos Prestados.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

st.write("")
st.write("")
st.write("")

st.write("📋 Tabela com as informações de logins e senhas de cada colaborador")
df_logins = ler_tabela(project_id="pagamento-terceirizado", 
                       dataset_id="pagamento_terceirizado", 
                       table_id="login_colaborador")
# st.dataframe(df_logins, hide_index=True)
# Exibe a tabela com HTML estilizado
st.markdown(df.to_html(index=False, escape=False), unsafe_allow_html=True)

st.write("")
st.write("")

st.write("📝 Informe abaixo os dados para ADICIONAR/EXCLUIR um novo login")

# Formulário de preenchimento do serviço prestado
with st.form(key="adicionar_excluir_login"):
    ACAO = st.selectbox(label="Informe se deseja ADICIONAR ou EXCLUIR o login", options=['ADICIONAR', 'EXCLUIR'])
    NEW_LOGIN = st.text_input(label="Informe o login")
    NOME_COMPLETO = st.text_input(label="Informe o nome completo do colaborador")
    input_buttom_submit_new_login = st.form_submit_button("Enviar")

if input_buttom_submit_new_login:
    if ACAO == 'ADICIONAR':
        if ((df_logins['LOGIN'] == NEW_LOGIN) & (df_logins['NOME_COMPLETO'] == NOME_COMPLETO)).any():
            st.warning("❌ Login já existente no banco de dados!")
        else:
            incluir_login(project_id="pagamento-terceirizado",
                        dataset_id="pagamento_terceirizado", 
                        table_id="login_colaborador",
                        LOGIN=NEW_LOGIN,
                        SENHA="123",
                        NOME_COMPLETO=NOME_COMPLETO)
            st.success("✅ Login incluído com sucesso!")
    elif ACAO == 'EXCLUIR':
        excluir_login(project_id="pagamento-terceirizado",
                      dataset_id="pagamento_terceirizado", 
                      table_id="login_colaborador",
                      LOGIN=NEW_LOGIN,
                      df_logins=df_logins)
        st.success("✅ Login excluído da base com sucesso!")