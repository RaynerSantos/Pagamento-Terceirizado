import os
import pandas as pd
import numpy as np
import streamlit as st
import gspread
from google.oauth2.service_account import Credentials
import json
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.styles.numbers import BUILTIN_FORMATS
from io import BytesIO
from datetime import datetime
from oauth2client.service_account import ServiceAccountCredentials
from utils import client, credentials

client = gspread.authorize(credentials)

# # Autenticação
# credentials = ServiceAccountCredentials.from_json_keyfile_name(json_path, scope)
# client = gspread.authorize(credentials)


# ===== Função para ler a planilha ===== #
def ler_tabela(sheet_name, worksheet_name):
    # Acessar a planilha
    spreadsheet = client.open(sheet_name)
    worksheet = spreadsheet.worksheet(worksheet_name)

    # Obter todos os valores da planilha
    data = worksheet.get_all_values()

    # Converter para um DataFrame do Pandas
    data = pd.DataFrame(data[1:], columns=data[0])  # A primeira linha vira cabeçalho
    return data

# Teste de leitura
df = ler_tabela(sheet_name="Pagamento_Terceirizado", worksheet_name="horas_colaborador")
print(df.head())

# ===== Função para incluir um novo serviço prestado ===== #
def incluir_servico(TERCEIRIZADO, SERVICO, DESCRICAO, PROJETO, PERIODO,
                    HORAS_TOTAIS, VALOR, PAGAMENTO_TOTAL, TIPO_COLABORADOR, QUEM_EMITE_A_NF, 
                    sheet_name, worksheet_name):
    # Acessar a planilha
    spreadsheet = client.open(sheet_name)
    worksheet = spreadsheet.worksheet(worksheet_name)

    nova_linha = [
        TERCEIRIZADO,
        SERVICO,
        DESCRICAO,
        PROJETO,
        PERIODO,
        HORAS_TOTAIS,
        VALOR,
        PAGAMENTO_TOTAL,
        TIPO_COLABORADOR,
        QUEM_EMITE_A_NF,
        datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ]
    
    # Adiciona ao final da worksheet
    worksheet.append_row(nova_linha, value_input_option='USER_ENTERED')
    print("✅ DataFrame atualizado com sucesso!")

# ===== Função para incluir um novo login ===== #
def incluir_login(sheet_name, worksheet_name, LOGIN, SENHA, NOME_COMPLETO):
    # Acessar a planilha
    spreadsheet = client.open(sheet_name)
    worksheet = spreadsheet.worksheet(worksheet_name)

    nova_linha = [
        LOGIN,
        SENHA,
        NOME_COMPLETO
    ]
    
    # Adiciona ao final da worksheet
    worksheet.append_row(nova_linha, value_input_option='USER_ENTERED')
    print("✅ Login incluído com sucesso!")
    return

# ===== Função para alterar a senha ===== #
def alterar_senha(sheet_name, worksheet_name, LOGIN, SENHA):
    # Ler a tabela de logins
    df_logins = ler_tabela(sheet_name=sheet_name, worksheet_name=worksheet_name)
    # Atualiza a senha
    df_logins.loc[df_logins["LOGIN"] == LOGIN, "SENHA"] = SENHA

     # Acessar a planilha
    spreadsheet = client.open(sheet_name)
    worksheet = spreadsheet.worksheet(worksheet_name)

    # Converter o DataFrame para uma lista de listas (formato aceito pelo gspread)
    data_atualizada = [df_logins.columns.values.tolist()] + df_logins.values.tolist()
    # Atualizar a planilha com os novos dados
    worksheet.update(data_atualizada)
    print("\n✅ Senha alterada com sucesso!")
    return

# ===== Função para excluir login ===== #
def excluir_login(sheet_name, worksheet_name, LOGIN):
    # Acessar a planilha
    spreadsheet = client.open(sheet_name)
    worksheet = spreadsheet.worksheet(worksheet_name)

    # Ler os dados
    data = worksheet.get_all_values()
    header = data[0]
    rows = data[1:]

    # Encontrar índice da linha com o LOGIN informado
    idx_para_remover = None
    for i, row in enumerate(rows):
        if row[0] == LOGIN:  # Considerando que LOGIN está na primeira coluna
            idx_para_remover = i + 2  # +2 para compensar cabeçalho e índice 1-based do Google Sheets
            break

    # Remover a linha, se encontrada
    if idx_para_remover:
        worksheet.delete_rows(idx_para_remover)
        print(f"✅ Login '{LOGIN}' excluído com sucesso!")
    else:
        print(f"⚠️ Login '{LOGIN}' não encontrado na planilha.")


###=== Função para excluir lançamento realizado ===###
def excluir_lancamento(sheet_name, worksheet_name, TERCEIRIZADO, PERIODO):
    # Acessar a planilha
    spreadsheet = client.open(sheet_name)
    worksheet = spreadsheet.worksheet(worksheet_name)

    # Obter os dados da planilha
    data = worksheet.get_all_values()
    header = data[0]
    rows = data[1:]

    # Identificar os índices das colunas relevantes
    try:
        idx_terceirizado = header.index("TERCEIRIZADO")
        idx_periodo = header.index("PERIODO")
    except ValueError as e:
        print("❌ Erro: Colunas 'TERCEIRIZADO' ou 'PERIODO' não encontradas.")
        return

    # Encontrar a linha que bate com os dois critérios
    linha_para_excluir = None
    for i, row in enumerate(rows):
        if (row[idx_terceirizado] == TERCEIRIZADO) and (row[idx_periodo] == PERIODO):
            linha_para_excluir = i + 2  # +2 por causa do cabeçalho e índice 1-based
            break

    # Excluir a linha, se encontrada
    if linha_para_excluir:
        worksheet.delete_rows(linha_para_excluir)
        print(f"✅ Lançamento de '{TERCEIRIZADO}' no período '{PERIODO}' foi excluído com sucesso!")
    else:
        print(f"⚠️ Nenhum lançamento encontrado para '{TERCEIRIZADO}' no período '{PERIODO}'.")


# incluir_servico(TERCEIRIZADO="Rayner", SERVICO="TESTE", DESCRICAO="TESTE", PROJETO="TESTE", PERIODO="TESTE", 
#                 HORAS_TOTAIS="12:00:00", VALOR="17", PAGAMENTO_TOTAL="134", 
#                     TIPO_COLABORADOR="TESTE", QUEM_EMITE_A_NF="TESTE",
#                     sheet_name="Pagamento_Terceirizado", worksheet_name="horas_colaborador")

# incluir_login(sheet_name="Pagamento_Terceirizado", worksheet_name="login_colaborador", 
#               LOGIN="YAN", SENHA="123!@", NOME_COMPLETO="YAN SABARENSE")

# alterar_senha(sheet_name="Pagamento_Terceirizado", worksheet_name="login_colaborador", LOGIN="rayner.santos", SENHA="123$")


# excluir_login(sheet_name="Pagamento_Terceirizado", worksheet_name="login_colaborador", LOGIN="YAN")

# excluir_lancamento(sheet_name="Pagamento_Terceirizado", worksheet_name="horas_colaborador", TERCEIRIZADO="Rayner", PERIODO="TESTE")




# Função para salvar a tabela em um único Excel com formatação
def salvar_excel_com_formatacao(bd):
    output = BytesIO()
    #=== Salvar em uma planilha em excel ===#
    # Crie uma nova planilha Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Pagamento"

    # # Remover as linhas de grade do Excel
    # ws.sheet_view.showGridLines = False

    # Define o estilo de preenchimento para o fundo do cabeçalho
    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Define o alinhamento centralizado sem quebra de texto
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

    # Define o alinhamento centralizado
    center_alignment = Alignment(horizontal="center", vertical="center")

    # Define o tamanho da fonte para todas as células
    font_size = Font(size=10)

    # Define a fonte em negrito
    bold_font = Font(bold=True, size=10)

    # Loop pelos DataFrames e escrevendo na planilha
    row_offset = 1  # Inicializa a contagem de linhas na planilha

    # Converter o DataFrame para linhas que o openpyxl pode usar
    rows = dataframe_to_rows(bd, index=False, header=True)

    # Define a borda fina
    borda_fina = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )

    bd["PAGAMENTO_TOTAL"] = pd.to_numeric(bd["PAGAMENTO_TOTAL"], errors="coerce")
    
    # Escreve os dados no Excel
    for j, row in enumerate(rows):
        ws.append(row)

        # Itera sobre todas as colunas e aplica centralização
        for col in range(1, len(row) + 1):  # Itera sobre as colunas
            cell = ws.cell(row=row_offset + j, column=col)
            cell.alignment = center_alignment  # Aplica alinhamento centralizado
            cell.font = font_size # Aplica o tamanho da fonte de 10 para todas as células
            cell.border = borda_fina  # adiciona borda a todas as células

            # Estiliza apenas o cabeçalho
            if j == 0:  # Cabeçalhos do multiíndice
                cell.fill = header_fill
                cell.font = bold_font
                cell.alignment = header_alignment # Estiliza o cabeçalho com quebra de texto

        # Formata como porcentagem colunas 6 e 7
        if (j > 0):  # Dados (exclui os cabeçalhos)
            for col in range(6, 9):  # Colunas 6 e 7
                cell = ws.cell(row=row_offset + j, column=col)
                if isinstance(cell.value, (int, float)):  # Verifica se é número
                    cell.number_format = 'R$ #,##0.00'  # Formato moeda brasileira (R$)
                    cell.value = float(cell.value)  # Converte o valor para floats
    
    # Salvar o Workbook no buffer
    wb.save(output)
    return output.getvalue()


