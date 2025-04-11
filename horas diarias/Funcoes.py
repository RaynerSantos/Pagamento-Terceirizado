import os
from google.cloud import bigquery
import pandas as pd
import numpy as np
import streamlit as st
import json
from google.oauth2 import service_account
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.styles.numbers import BUILTIN_FORMATS
from io import BytesIO
from datetime import datetime

os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = "C:\PROJETOS\Pagamento Terceirizado\Ignorar\pagamento-terceirizado-467d410b51b5.json"

# Nome do projeto, dataset e tabela
project_id = "pagamento-terceirizado"
dataset_id = "pagamento_terceirizado"
table_id = "login_colaborador"  # "horas_diaria_colaborador" / "login_colaborador"


# ===== Função para ler a tabela ===== #
def ler_tabela(project_id, dataset_id, table_id):
    # Inicializa o cliente
    client = bigquery.Client(project=project_id)

    # Consulta SQL para ler a tabela
    query = f"SELECT * FROM `{project_id}.{dataset_id}.{table_id}`"
    df = client.query(query).to_dataframe()

    return df

# ===== Função para incluir um novo serviço prestado ===== #
def incluir_horas(project_id, dataset_id, table_id, 
                  PROJETO, NOME_COMPLETO, CPF, TELEFONE, VALOR_HORA, DATA, QTD_HORAS, BANCO, AGENCIA, CONTA, TIPO_PIX, CHAVE_PIX):
    # Inicializa o cliente
    client = bigquery.Client(project=project_id)

    # DataFrame com dados novos
    novos_dados = pd.DataFrame({
        "PROJETO": [PROJETO],
        "NOME_COMPLETO": [NOME_COMPLETO],
        "CPF": [CPF],
        "TELEFONE": [TELEFONE],
        "VALOR_HORA": [VALOR_HORA],
        "DATA": [DATA],
        "QTD_HORAS": [QTD_HORAS],
        "VALOR_TOTAL": [round(VALOR_HORA * QTD_HORAS, 2)],
        "BANCO": [BANCO],
        "AGENCIA": [AGENCIA],
        "CONTA": [CONTA],
        "TIPO_PIX": [TIPO_PIX],
        "CHAVE_PIX": [CHAVE_PIX],
        "DATA_LANCAMENTO": [datetime.now()]
    })

    # Define a tabela completa
    tabela_destino = f"{project_id}.{dataset_id}.{table_id}"

    # Insere no modo append
    job = client.load_table_from_dataframe(novos_dados, tabela_destino)
    job.result()  # Espera o job terminar
    print("Dados inseridos com sucesso!")
    return


# Função para salvar a tabela em um único Excel com formatação
def salvar_excel_com_formatacao(bd):
    output = BytesIO()
    #=== Salvar em uma planilha em excel ===#
    # Crie uma nova planilha Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Pagamento"

    # # Remover as linhas de grade do Excel
    # ws.sheet_view.showGridLines = False

    # Define o estilo de preenchimento para o fundo do cabeçalho
    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Define o alinhamento centralizado sem quebra de texto
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

    # Define o alinhamento centralizado
    center_alignment = Alignment(horizontal="center", vertical="center")

    # Define o tamanho da fonte para todas as células
    font_size = Font(size=10)

    # Define a fonte em negrito
    bold_font = Font(bold=True, size=10)

    # Loop pelos DataFrames e escrevendo na planilha
    row_offset = 1  # Inicializa a contagem de linhas na planilha

    # Converter o DataFrame para linhas que o openpyxl pode usar
    rows = dataframe_to_rows(bd, index=False, header=True)

    # Define a borda fina
    borda_fina = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )
    
    # Escreve os dados no Excel
    for j, row in enumerate(rows):
        ws.append(row)

        # Itera sobre todas as colunas e aplica centralização
        for col in range(1, len(row) + 1):  # Itera sobre as colunas
            cell = ws.cell(row=row_offset + j, column=col)
            cell.alignment = center_alignment  # Aplica alinhamento centralizado
            cell.font = font_size # Aplica o tamanho da fonte de 10 para todas as células
            cell.border = borda_fina  # adiciona borda a todas as células

            # Estiliza apenas o cabeçalho
            if j == 0:  # Cabeçalhos do multiíndice
                cell.fill = header_fill
                cell.font = bold_font
                cell.alignment = header_alignment # Estiliza o cabeçalho com quebra de texto

        # Formata como porcentagem colunas 6 e 7
        if (j > 0):  # Dados (exclui os cabeçalhos)
            for col in range(6, 9):  # Colunas 6 e 7
                cell = ws.cell(row=row_offset + j, column=col)
                if isinstance(cell.value, (int, float)):  # Verifica se é número
                    cell.number_format = 'R$ #,##0.00'  # Formato moeda brasileira (R$)
                    cell.value = float(cell.value)  # Converte o valor para floats
    
    # Salvar o Workbook no buffer
    wb.save(output)
    return output.getvalue()